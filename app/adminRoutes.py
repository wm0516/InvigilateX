import os
import re
import warnings
from io import BytesIO
from collections import defaultdict
from datetime import datetime, time, date
import random
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import NamedStyle
import PyPDF2
from flask import render_template, request, redirect, url_for,flash, session, jsonify, send_file
from flask_bcrypt import Bcrypt
from itsdangerous import URLSafeTimedSerializer
from sqlalchemy import func, and_, or_, case, desc
from app import app
from .authRoutes import login_required
from .backend import *
from .database import *

# Initialize serializer and bcrypt
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
bcrypt = Bcrypt()

# Upload configuration
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)



# -------------------------------
# Function handle file upload
# -------------------------------
def handle_file_upload(file_key, expected_cols, process_row_fn, redirect_endpoint, usecols="A:Z", skiprows=1):
    file = request.files.get(file_key)
    if file and file.filename:
        try:
            file_stream = BytesIO(file.read())
            excel_file = pd.ExcelFile(file_stream)
            records_added = records_failed = 0

            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        usecols=usecols,
                        skiprows=skiprows,      # skip the first row(s)
                        header=0,               # take the next row as header
                        dtype=str
                    )
                    df.columns = [str(col).strip().lower() for col in df.columns]

                    if not set(expected_cols).issubset(df.columns):
                        flash(f"{df.columns}","error")
                        raise ValueError(f"Excel columns do not match expected format: {df.columns.tolist()}")

                    # normalize all columns
                    for col in df.columns:
                        df[col] = df[col].apply(lambda x: str(x).strip() if isinstance(x, str) else x)

                    # process each row using the provided callback
                    for _, row in df.iterrows():
                        try:
                            success, message = process_row_fn(row)
                            if success:
                                records_added += 1
                            else:
                                records_failed += 1
                                flash(f"Row failed: {message}", "error")
                        except Exception as row_err:
                            records_failed += 1
                            flash(f"Row processing error: {str(row_err)} | Data: {row.to_dict()}", "error")

                except Exception as sheet_err:
                    pass
            if records_added > 0:
                flash(f"Successfully uploaded {records_added} record(s)", "success")
            if records_failed > 0:
                flash(f"Failed to upload {records_failed} record(s)", "error")
            if records_added == 0 and records_failed == 0:
                flash("No data uploaded", "error")

            return redirect(url_for(redirect_endpoint))
        except Exception as e:
            flash("File processing error: File upload in wrong format", "error")
            return redirect(url_for(redirect_endpoint))
    else:
        flash("No file uploaded", "error")
        return redirect(url_for(redirect_endpoint))


# -------------------------------
# Function for Admin ManageCourse Download Overall Excel File Format 
# -------------------------------
def generate_managecourse_template():
    warnings.simplefilter("ignore", UserWarning)
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None, "Workbook has no active worksheet"
    ws.title = "Courses"
    
    # --- Header row ---
    ws.append([])  # First row empty
    headers = ['Program Code', 'Course Code', 'Course Section', 'Course Intake', 'Course Name', 'Credit Hour', 'No. of Students']
    ws.append(headers)

    # --- Hidden sheet for dropdown lists ---
    ws_lists = wb.create_sheet(title="Lists")

    # --- Department Codes ---
    departments = [d.departmentCode for d in Department.query.order_by(Department.departmentCode).all()]
    for i, dept in enumerate(departments, start=1):
        ws_lists[f"A{i}"] = dept

    # --- Intake Semester/Month-Year Combos ---
    current_year = datetime.now().year
    semesters_by_year = {
        current_year - 1: ['AUG', 'OCT'],
        current_year: ['JAN', 'APR', 'AUG', 'OCT'],
        current_year + 1: ['JAN', 'APR']
    }

    # Flatten into a single list like "JAN2026", "APR2026", etc.
    intake_list = []
    for year, months in semesters_by_year.items():
        for month in months:
            intake_list.append(f"{month}{year}")

    # Write intake options into hidden sheet column B
    for i, intake in enumerate(intake_list, start=1):
        ws_lists[f"B{i}"] = intake

    # Hide the Lists sheet
    ws_lists.sheet_state = 'hidden'

    # --- Data Validations ---
    if departments:
        dv_dept = DataValidation(type="list", formula1=f"=Lists!$A$1:$A${len(departments)}", allow_blank=False)
        ws.add_data_validation(dv_dept)
        dv_dept.add("A3:A1000")  # Department Code column

    if intake_list:
        dv_intake = DataValidation(type="list", formula1=f"=Lists!$B$1:$B${len(intake_list)}", allow_blank=False)
        ws.add_data_validation(dv_intake)
        dv_intake.add("D3:D1000")  # Course Intake column

    # Return as BytesIO for Flask send_file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# -------------------------------
# Function for Admin ManageCourse Download Excel File Template  
# -------------------------------
@app.route('/download_course_template')
@login_required
def download_course_template():
    output = generate_managecourse_template()
    return send_file(
        output,
        as_attachment=True, 
        download_name=f"ManageCourse.xlsx", # type: ignore[arg-type]
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -------------------------------
# Function for Admin ManageCourse Route Upload File
# -------------------------------
def process_course_row(row):
    user_id = session.get('user_id')
    return create_course_and_exam(
        userid      = user_id,
        department  = str(row['program code']).strip(),
        code        = str(row['course code']).strip().replace(" ", ""),
        section     = str(row['course section']).strip().replace(" ", ""),
        name        = str(row['course name']).strip(),
        hour        = int(row['credit hour']),
        students    = int(row['no. of students']),
        intake      = str(row['course intake']).strip(),
    )

# -------------------------------
# Read All LecturerName Under The Selected Department For ManageCoursePage
# -------------------------------   
@app.route('/get_lecturers_by_department/<department_code>')
@login_required
def get_lecturers_by_department(department_code):
    # Filter userDepartment and multiple userLevel values
    lecturers = User.query.filter(User.userDepartment == department_code,User.userLevel.in_([1, 2, 3, 4])).all()
    lecturers_list = [{"userId": l.userId, "userName": l.userName} for l in lecturers]
    return jsonify(lecturers_list)

# -------------------------------
# Read All CourseCodeSection and Return all the selected data details Under The ManageCourseEditPage
# -------------------------------
@app.route('/get_courseCodeSection/<path:courseCodeSection_select>')
@login_required
def get_courseCodeSection(courseCodeSection_select):
    course = Course.query.filter_by(courseCodeSectionIntake=courseCodeSection_select).first()
    if not course:
        return jsonify({"error": "Course not found"}), 404
    
    return jsonify({
        "courseDepartment": course.courseDepartment,
        "courseName": course.courseName,
        "courseHour": course.courseHour,
        "courseStudent": course.courseStudent,
        "courseStatus": course.courseStatus,
        "practicalLecturerId": course.coursePractical,
        "tutorialLecturerId": course.courseTutorial,
        "classLecturerId": course.courseLecturer
    })

# -------------------------------
# Function for Admin ManageCourse Route
# -------------------------------
@app.route('/admin/manageCourse', methods=['GET', 'POST'])
@login_required
def admin_manageCourse():
    user_id = session.get('user_id')

    # === Get all courses, optionally filtered by intake ===
    course_query = Course.query.order_by(
        Course.courseStatus.desc(),
        Course.courseCodeSectionIntake.asc()
    )

    course_data = course_query.all()
    department_data = Department.query.all()
    course_id = request.form.get('editCourseSelect')
    course_select = Course.query.filter_by(courseCodeSectionIntake=course_id).first()

    # Count rows with missing/empty values
    error_rows = Course.query.filter(
        (Course.courseDepartment.is_(None)) | (Course.courseDepartment == '')
    ).count()

    # Courses by department safely (only active courses in selected intake)
    courses_by_department_raw = (
        db.session.query(
            func.coalesce(Department.departmentCode, "Unknown").label("department"),
            func.count(Course.courseCodeSectionIntake).label("course_count")
        )
        .outerjoin(Course, Department.departmentCode == Course.courseDepartment)
        .filter(Course.courseStatus == True)
        .group_by(func.coalesce(Department.departmentCode, "Unknown"))
        .having(func.count(Course.courseCodeSectionIntake) > 0)
        .order_by(func.coalesce(Department.departmentCode, "Unknown").asc())
        .all()
        or []
    )

    courses_by_department = [
        {"department": dept_code, "count": count}
        for dept_code, count in courses_by_department_raw
    ]

    # === POST Handling ===
    if request.method == 'POST':
        form_type = request.form.get('form_type')

        # --- Upload Section ---
        if form_type == 'upload':
            return handle_file_upload(
                file_key='course_file',
                expected_cols=['program code', 'course code', 'course section', 'course intake', 'course name', 'credit hour', 'no. of students'],
                process_row_fn=process_course_row,
                redirect_endpoint='admin_manageCourse',
                usecols="A:G",
                skiprows=1 
            )

        # --- Edit Section ---
        elif form_type == 'edit':
            action = request.form.get('action')
            code = request.form.get('courseCode', '').strip()
            section = request.form.get('courseSection', '').strip()
            intake = request.form.get('intakeSemesterEdit', '').strip()  # fixed name
            courseCodeSection_text = f"{intake}/{code}/{section}".upper() if code and section and intake else None

            if action == 'update' and course_select:
                course_status = True if request.form.get('courseStatus') == '1' else False

                # --- UPDATE COURSE ---
                if course_status:
                    # Validate PK uniqueness before updating
                    if courseCodeSection_text != course_select.courseCodeSectionIntake:
                        existing_course = Course.query.get(courseCodeSection_text)
                        if existing_course:
                            flash("This Course Code/Section/Intake already exists!", "error")
                            return redirect(request.url)

                    # Update PK and other fields
                    course_select.courseCodeSectionIntake = courseCodeSection_text
                    course_select.courseDepartment  = request.form.get('departmentCode', '').strip()
                    course_select.courseName        = request.form.get('courseName', '').strip()
                    course_select.coursePractical   = request.form.get('practicalLecturerSelect', '').strip() or None
                    course_select.courseTutorial    = request.form.get('tutorialLecturerSelect', '').strip() or None
                    course_select.courseLecturer    = request.form.get('lecturerSelect', '').strip() or None
                    course_select.courseStatus      = True

                    # Safe integer conversion
                    try:
                        course_select.courseHour = int(request.form.get('courseHour', 0))
                    except (ValueError, TypeError):
                        course_select.courseHour = 0

                    try:
                        course_select.courseStudent = int(request.form.get('courseStudent', 0))
                    except (ValueError, TypeError):
                        course_select.courseStudent = 0

                    # Validate required fields
                    required_fields = [
                        course_select.courseDepartment,
                        course_select.courseName,
                        course_select.coursePractical,
                        course_select.courseTutorial,
                        course_select.courseLecturer,
                        course_select.courseHour,
                        course_select.courseStudent
                    ]

                    if all(f not in (None, '') for f in required_fields):
                        # Ensure Exam exists
                        if not course_select.courseExamId:
                            new_exam = Exam(examOutput=None)
                            db.session.add(new_exam)
                            db.session.flush()  # Assign examId
                            course_select.courseExamId = new_exam.examId

                    db.session.commit()
                    flash(f"Course [{course_select.courseCodeSectionIntake} - {course_select.courseName}] updated successfully", "success")
                    record_action("EDIT COURSE", "COURSE", course_select.courseCodeSectionIntake, user_id)

                # --- DEACTIVATE COURSE ---
                else:
                    course_select.courseStatus = False
                    db.session.commit()
                    flash(f"Course [{course_select.courseCodeSectionIntake} - {course_select.courseName}] deactivated successfully", "success")
                    record_action("DEACTIVATE COURSE", "COURSE", course_select.courseCodeSectionIntake, user_id)

                return redirect(url_for('admin_manageCourse'))


        # --- Manual Add Section ---
        elif form_type == 'manual':
            def safe_int(value, default=0):
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return default

            form_data = {
                "userid"    : user_id,
                "department": request.form.get('departmentCode', '').strip(),
                "code"      : request.form.get('courseCode', '').replace(' ', ''),
                "section"   : request.form.get('courseSection', '').replace(' ', ''),
                "name"      : request.form.get('courseName', '').strip(),
                "hour"      : safe_int(request.form.get('courseHour')),
                "students"  : safe_int(request.form.get('courseStudent')),
                "intake"    : request.form.get('intakeSemester', '').strip(),
            }
            success, message = create_course_and_exam(**form_data)
            flash(message, "success" if success else "error")
            return redirect(url_for('admin_manageCourse'))

    # === GET Request ===
    return render_template('admin/adminManageCourse.html', active_tab='admin_manageCoursetab', course_data=course_data, course_select=course_select,
                            department_data=department_data, courses_by_department=courses_by_department, error_rows=error_rows)





# -------------------------------
# Get Department Details for ManageDepartmentEditPage
# -------------------------------
@app.route('/get_department/<path:department_code>')
@login_required
def get_department(department_code):
    dept = Department.query.filter_by(departmentCode=department_code).first()
    if not dept:
        return jsonify({"error": "Department not found"}), 404
    return jsonify({
        "departmentCode": dept.departmentCode,
        "departmentName": dept.departmentName,
        "deanId"        : dept.deanId or None,
        "hosId"         : dept.hosId or None,
        "hopId"         : dept.hopId or None
    })

def validate_user_role(user_id, expected_level, department_code, role_name):
    if not user_id:
        return None

    user = User.query.filter_by(userId=user_id, userLevel=expected_level).first()
    if not user or user.userDepartment != department_code:
        flash(f"Selected {role_name} does not belong to this department. Ignoring {role_name} selection.", "error")
        return None

    return user_id

# -------------------------------
# Admin Manage Department
# -------------------------------
@app.route('/admin/manageDepartment', methods=['GET', 'POST'])
@login_required
def admin_manageDepartment():
    user_id = session.get('user_id')
    # Load all departments and stats
    department_data     = Department.query.all()
    total_department    = len(department_data)
    deans               = User.query.filter_by(userLevel="DEAN").all()
    hoss                = User.query.filter_by(userLevel="HOS").all()
    hops                = User.query.filter_by(userLevel="HOP").all()

    # For edit section
    department_selected_code = request.form.get('editDepartment')
    department_select        = Department.query.filter_by(departmentCode=department_selected_code).first()

    if request.method == 'POST':
        form_type = request.form.get('form_type')

        # ---------------- Manual Section ----------------
        if form_type == 'manual':
            departmentCode = request.form.get('departmentCode', '').strip().upper()
            departmentName = request.form.get('departmentName', '').strip().upper()

            # Check if department code already exists
            if Department.query.filter_by(departmentCode=departmentCode).first():
                flash("Department code already exists. Please use a unique code.", "error")
            else:
                db.session.add(
                    Department(
                        departmentCode=departmentCode, 
                        departmentName=departmentName,
                        )
                    )
                db.session.commit()
                flash(f"New department [{departmentCode}] added", "success")               
                record_action("ADD NEW DEPARTMENT", "DEPARTMENT", departmentCode, user_id)
                return redirect(url_for('admin_manageDepartment'))

        # ---------------- Edit Section ----------------
        elif form_type == 'edit' and department_select:
            action           = request.form.get('action')
            departmentName   = request.form.get('departmentName')
            hosId            = request.form.get('hosName') or None
            deanId           = request.form.get('deanName') or None
            hopId            = request.form.get('hopName') or None

            deanId = validate_user_role(deanId, "DEAN", department_select.departmentCode, "Dean")
            hosId  = validate_user_role(hosId, "HOS", department_select.departmentCode, "HOS")
            hopId  = validate_user_role(hopId, "HOP", department_select.departmentCode, "HOP")

            if action == 'update':
                department_select.departmentName = departmentName

                # These can be None → will reset previous values
                department_select.deanId = deanId
                department_select.hosId  = hosId
                department_select.hopId  = hopId
                db.session.commit()
                flash(f"Department [{department_select.departmentCode}] updated successfully", "success")
                record_action("EDIT DEPARTMENT", "DEPARTMENT", department_select.departmentCode, user_id)
            
            elif action == 'delete':
                users_using_department = User.query.filter_by(userDepartment=department_select.departmentCode).count()
                if users_using_department > 0:
                    flash(f"Cannot delete department. There are number of {users_using_department} users still assigned to this department.", "error")
                else:
                    db.session.delete(department_select)
                    db.session.commit()
                    flash(f"Department [{department_select.departmentCode}] deleted successfully", "success")
                    record_action("DELETE DEPARTMENT", "DEPARTMENT", department_select.departmentCode, user_id)

            return redirect(url_for('admin_manageDepartment'))
    return render_template('admin/adminManageDepartment.html', active_tab='admin_manageDepartmenttab', department_data=department_data, department_select=department_select, total_department=total_department, deans=deans, hoss=hoss, hops=hops)




# -------------------------------
# Get Venue Details for ManageVenueEditPage
# -------------------------------
@app.route('/get_venue/<venue_number>')
@login_required
def get_venue(venue_number):
    venue = Venue.query.filter_by(venueNumber=venue_number).first()
    if not venue:
        return jsonify({"error": "Venue not found"}), 404

    return jsonify({
        "venueNumber"   : venue.venueNumber,
        "venueLevel"    : venue.venueLevel,
        "venueCapacity" : venue.venueCapacity
    })

# -------------------------------
# Function for Admin ManageVenue Route
# -------------------------------
@app.route('/admin/manageVenue', methods=['GET', 'POST'])
@login_required
def admin_manageVenue():
    user_id = session.get('user_id')
    # Load all venues and stats
    venue_data = Venue.query.order_by(Venue.venueLevel.asc()).all()

    # For edit section
    venue_selected_number   = request.form.get('editVenueNumber')
    venue_select            = Venue.query.filter_by(venueNumber=venue_selected_number).first()

    if request.method == 'POST':
        form_type = request.form.get('form_type')

        # ---------------- Manual Section ----------------
        if form_type == 'manual':
            venueNumber     = request.form.get('venueNumber', '').strip().upper()
            venueLevel      = request.form.get('venueLevel', '').strip()
            venueCapacity   = request.form.get('venueCapacity', '').strip()

            # Check if venue already exists
            if Venue.query.filter_by(venueNumber=venueNumber).first():
                flash("Venue Room Already Exists", "error")
            else:
                try: 
                    capacity = int(venueCapacity)
                    if capacity < 0:
                        raise ValueError
                    db.session.add(Venue(
                        venueNumber=venueNumber,
                        venueLevel=venueLevel,
                        venueCapacity=capacity
                    ))
                    db.session.commit()
                    flash(f"Venue [{venueNumber}] Added", "success")
                    record_action("ADD NEW VENUE", "VENUE", venueNumber, user_id)
                except ValueError:
                    flash("Capacity must be a non-negative integer", "error")
            return redirect(url_for('admin_manageVenue'))

        # ---------------- Edit Section ----------------
        elif form_type == 'edit' and venue_select:
            action          = request.form.get('action')
            venueLevel      = request.form.get('venueLevel')
            venueCapacity   = request.form.get('venueCapacity')

            if action == 'update':
                try:
                    capacity = int(venueCapacity)
                    if capacity < 0:
                        raise ValueError
                    venue_select.venueLevel = venueLevel
                    venue_select.venueCapacity = capacity
                    db.session.commit()
                    flash(f"Venue [{venue_select.venueNumber}] Updated", "success")
                    record_action("EDIT VENUE", "VENUE", venue_select.venueNumber, user_id)
                except ValueError:
                    flash("Capacity must be a non-negative integer", "error")

            elif action == 'delete':
                try:
                    venue_number = venue_select.venueNumber
                    exam_count = VenueSession.query.filter_by(venueNumber=venue_number).count()

                    if exam_count > 0:
                        flash(f"Cannot delete venue. It is still used by {exam_count} exams", "error")
                    else:
                        db.session.delete(venue_select)
                        db.session.commit()
                        flash(f"Venue [{venue_select.venueNumber}] deleted successfully", "success")
                        record_action("DELETE VENUE", "VENUE", venue_select.venueNumber, user_id)
                except Exception as e:
                    db.session.rollback()
                    flash(f"Failed to delete venue: {str(e)}", "error")
            return redirect(url_for('admin_manageVenue'))
    # Render template
    return render_template('admin/adminManageVenue.html', active_tab='admin_manageVenuetab', venue_data=venue_data, venue_select=venue_select)






# -------------------------------
# Function for Admin ManageExam Download Excel File Format
# -------------------------------
def generate_manageexam_template():
    warnings.simplefilter("ignore", UserWarning)
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None, "Workbook has no active worksheet"
    ws.title = "Exams"

    # --- Row 1 empty
    ws.append([])

    # --- Row 2 headers (updated)
    headers = ['Exam Date','Day','Start','End','Program','Course Code/Section','Course Name','Lecturer','Total Student by venue','Venue']
    ws.append(headers)

    # === Formatting for date/time ===
    date_style = NamedStyle(name="exam_date", number_format="MM/DD/YYYY")
    time_style = NamedStyle(name="exam_time", number_format="hh:mm:ss AM/PM")

    # Add styles ONCE
    if "exam_date" not in wb.named_styles:
        wb.add_named_style(date_style)
    if "exam_time" not in wb.named_styles:
        wb.add_named_style(time_style)

    for row in range(3, 503):
        ws[f"A{row}"].style = date_style
        ws[f"B{row}"] = f'=IF(A{row}="","",TEXT(A{row},"dddd"))'
        ws[f"C{row}"].style = time_style  # Start Time
        ws[f"D{row}"].style = time_style  # End Time

    # === Hidden sheet for lookup lists ===
    ws_lists = wb.create_sheet(title="Lists")

    # --- Fetch courses and exams ---
    courses = (
        db.session.query(Course, Exam)
        .join(Exam, Course.courseExamId == Exam.examId).filter(
            and_(Course.courseStatus == True)).all()
    )

    # --- Fill lookup data ---
    for i, (c, e) in enumerate(courses, start=1):
        lecturer_names = []
        if c.practicalLecturer:
            lecturer_names.append(c.practicalLecturer.userName)
        if c.tutorialLecturer:
            lecturer_names.append(c.tutorialLecturer.userName)
        if c.classLecturer:
            lecturer_names.append(c.classLecturer.userName)

        lecturer_str = ", ".join(lecturer_names) if lecturer_names else "None"
        ws_lists[f"A{i}"] = c.courseCodeSectionIntake
        ws_lists[f"B{i}"] = c.courseDepartment
        ws_lists[f"C{i}"] = c.courseName
        ws_lists[f"D{i}"] = lecturer_str
        ws_lists[f"E{i}"] = c.courseStudent

    # --- Venues ---
    venues = Venue.query.all()
    for i, v in enumerate(venues, start=1):
        ws_lists[f"F{i}"] = v.venueNumber

    # === Data Validations ===
    if courses:
        # Dropdown for Course Code
        dv_course = DataValidation(type="list",formula1=f"=Lists!$A$1:$A${len(courses)}",allow_blank=False)
        ws.add_data_validation(dv_course)
        dv_course.add("F3:F502")

        # Auto-fill Course Name and Total number of students
        for row in range(3, 503):
            ws[f"E{row}"] = (f'=IF(F{row}="","",'f'VLOOKUP(F{row},Lists!$A$1:$D${len(courses)},2,FALSE))')
            ws[f"G{row}"] = (f'=IF(F{row}="","",'f'VLOOKUP(F{row},Lists!$A$1:$D${len(courses)},3,FALSE))')
            ws[f"H{row}"] = (f'=IF(F{row}="","",'f'VLOOKUP(F{row},Lists!$A$1:$D${len(courses)},4,FALSE))')
            ws[f"I{row}"] = f'=IF(F{row}="","",VLOOKUP(F{row},Lists!$A$1:$E${len(courses)},5,FALSE))'

    # Dropdown for Exam Venue
    if venues:
        dv_venue = DataValidation(type="list",formula1=f"=Lists!$F$1:$F${len(venues)}",allow_blank=False)
        ws.add_data_validation(dv_venue)
        dv_venue.add("J3:J502")

    # Hide lookup sheet
    ws_lists.sheet_state = 'hidden'
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# -------------------------------
# Function for Admin ManageExam Download Excel File Template
# -------------------------------
@app.route('/download_exam_template')
@login_required
def download_exam_template():
    output = generate_manageexam_template()
    return send_file(
        output,
        as_attachment=True,
        download_name="ManageExam.xlsx", # type: ignore[arg-type]
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -------------------------------
# Function for Admin ManageExam Route Upload File Combine Date and Time
# -------------------------------
def process_exam_row(row, slot_share_dt, slot_open_dt):
    user_id = session.get('user_id')

    important_columns = ['exam date', 'start', 'end', 'course code/section', 'total student by venue', 'venue']
    if all(not row.get(col) or str(row.get(col)).strip() == '' for col in important_columns):
        return None, None 

    # Parse exam start & end datetimes
    try:
        start_dt = parse_datetime(row.get('exam date'), row.get('start'))
        end_dt   = parse_datetime(row.get('exam date'), row.get('end'))
    except ValueError as e:
        return False, str(e)

    # Handle overnight exams
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)

    # Parse total students
    try:
        requested_capacity = int(row.get('total student by venue', 0))
    except ValueError:
        return False, f"Invalid total number of students by venue: {row.get('total student by venue')}"

    # Parse venue
    venue = str(row.get('venue', '')).strip().upper()
    if not venue:
        return False, "Exam Venue missing"

    venue_obj = Venue.query.filter_by(venueNumber=venue).first()
    if not venue_obj:
        return False, f"Venue {venue} not found in database"

    # Create exam & related records
    success, msg = create_exam_and_related(user_id, start_dt, end_dt, str(row.get('course code/section')).strip().upper(), [venue], [requested_capacity], slot_share_dt, slot_open_dt)
    if not success:
        flash(msg, "error")
        return False, msg

    return True, None


# -------------------------------
# Get ExamDetails for ManageExamEditPage
# -------------------------------
@app.route('/get_exam_details/<path:course_code>')
@login_required
def get_exam_details(course_code):
    course = Course.query.filter_by(courseCodeSectionIntake=course_code).first()

    if not course:
        return jsonify({"error": "Course not found"}), 404
    exam = course.exam
    if not exam:
        return jsonify({"error": "Exam not found"}), 404

    # Latest attendance record (optional)
    attendance = (
        InvigilatorAttendance.query
        .join(InvigilationReport)
        .filter(InvigilationReport.examId == exam.examId)
        .order_by(InvigilatorAttendance.timeExpire.desc())
        .first()
    )

    exam_venues = []
    for ve in exam.venue_availabilities:
        session = ve.session
        venue = session.venue
        exam_venues.append({
            "venueNumber": venue.venueNumber,
            "venueLevel": venue.venueLevel,
            "venueCapacity": venue.venueCapacity,
            "studentCount": ve.studentCount,
            "startDateTime": session.startDateTime.strftime("%Y-%m-%dT%H:%M"),
            "endDateTime": session.endDateTime.strftime("%Y-%m-%dT%H:%M")
        })

    # Use first venue session for exam times
    examTimeCreate = exam_venues[0]["startDateTime"] if exam_venues else ""
    examTimeExpire = exam_venues[0]["endDateTime"] if exam_venues else ""

    response_data = {
        "courseCode": course.courseCodeSectionIntake,
        "courseName": course.courseName,
        "courseDepartment": course.courseDepartment,
        "totalStudents": course.courseStudent,
        "examId": exam.examId,
        "examStatus": exam.examStatus,
        "examVenues": exam_venues,
        "examTimeCreate": examTimeCreate,
        "examTimeExpire": examTimeExpire
    }

    return jsonify(response_data)

# -------------------------------
# Reformat the datetime for ManageExamEditPage
# -------------------------------
def parse_datetime(date_str, time_str):
    if not date_str or not time_str:
        raise ValueError(f"Missing date or time: {date_str}, {time_str}")

    # Try to parse date
    date_formats = ["%m/%d/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"]
    for fmt in date_formats:
        try:
            if isinstance(date_str, datetime):
                exam_date = date_str.date()
            else:
                exam_date = datetime.strptime(str(date_str).strip(), fmt).date()
            break
        except ValueError:
            continue
    else:
        raise ValueError(f"Invalid date format: {date_str}")

    # Parse time
    if isinstance(time_str, time):
        exam_time = time_str
    elif isinstance(time_str, datetime):
        exam_time = time_str.time()
    elif isinstance(time_str, str):
        exam_time = parse_excel_time(time_str)
    else:
        raise ValueError(f"Invalid time value: {time_str}")

    return datetime.combine(exam_date, exam_time)


def parse_excel_time(val):
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%I:%M:%S %p", "%H:%M:%S"):
            try:
                return datetime.strptime(val, fmt).time()
            except ValueError:
                pass
    raise ValueError(f"Invalid time format: {val}")

def safe_iso(val, field):
    if not val:
        raise ValueError(f"{field} is missing")
    return datetime.fromisoformat(val)


# Remove all venue exams, attendances, reports, and rollback pending hours SAFELY.
def reset_exam_relations(exam):
    if not exam:
        return

    # 1. Calculate total exam duration (from venue sessions)
    total_hours = 0.0
    for ve in exam.venue_availabilities:
        session = ve.session
        if session and session.startDateTime and session.endDateTime:
            duration = (session.endDateTime - session.startDateTime).total_seconds() / 3600.0
            total_hours += duration

    # 2. Roll back invigilator hours + delete reports
    reports = InvigilationReport.query.filter_by(examId=exam.examId).all()

    for report in reports:
        for att in report.attendances:
            inv = att.invigilator
            if inv and total_hours > 0:
                inv.userPendingCumulativeHours = max(0.0, (inv.userPendingCumulativeHours or 0.0) - total_hours)
        db.session.delete(report)

    # 3. Remove venue-exam relations safely
    for ve in exam.venue_availabilities:
        db.session.delete(ve)

    # 4. Reset exam fields
    exam.examOutput = None
    db.session.commit()



# -------------------------------
# Reassign invigilator for ManageExamEditPage
# -------------------------------
def adjust_exam(exam, new_start, new_end, new_venues, new_students, time_open, time_expire):
    user_id = session.get("user_id")

    # Validate input
    if not new_start or not new_end:
        raise ValueError("Exam start and end time required")
    if time_open is None or time_expire is None:
        raise ValueError("Open / Expire time cannot be NULL")

    # FULL RESET (important)
    reset_exam_relations(exam)

    # RECREATE using SAME logic as upload
    success, msg = create_exam_and_related(
        user=user_id,
        start_dt=new_start,
        end_dt=new_end,
        courseSection=exam.course.courseCodeSectionIntake,
        venue_list=new_venues,
        studentPerVenue_list=new_students,
        open=time_open,
        close=time_expire
    )

    if not success:
        db.session.rollback()
        raise ValueError(msg)
    db.session.commit()

# -------------------------------
# Function for Admin ManageExam Route
# -------------------------------
@app.route('/admin/manageExam', methods=['GET', 'POST'])
@login_required
def admin_manageExam():
    user_id = session.get('user_id')

    # Load static data
    department_data = Department.query.all()
    venue_data = Venue.query.order_by(Venue.venueCapacity.asc()).all()

    # Main exam list (ordered correctly)
    exam_data = (
        Exam.query
        .join(Exam.course)
        .outerjoin(Exam.venue_availabilities)
        .outerjoin(VenueExam.session)
        .filter(Course.courseStatus == True)
        .group_by(Exam.examId)
        .order_by(
            Exam.examStatus.desc(),
            func.min(VenueSession.startDateTime).is_(None).asc(),  # unscheduled first
            func.min(VenueSession.startDateTime).asc(),           # earliest exam first
            Exam.examId.asc(),
        )
        .all()
    )

    display_exam_data = [e for e in exam_data if e.examStatus]

    # Counters
    unassigned_exam = len([
        e for e in exam_data
        if e.examStatus and not e.venue_availabilities
    ])

    total_exam_activated = (
        Exam.query
        .filter(
            Exam.examStatus == True,
            Exam.venue_availabilities.any()
        )
        .count()
    )

    # For edit section
    edit_exam_data = (
        Exam.query
        .join(Exam.course)
        .filter(
            Exam.examStatus == True,
            Course.courseStatus == True
        )
        .order_by(Course.courseCodeSectionIntake.asc())
        .all()
    )

    exam_selected_code = request.form.get('editExamCourseSection')
    course = (
        Course.query
        .filter(Course.courseCodeSectionIntake == exam_selected_code)
        .first()
    )
    exam_select = course.exam if course else None

    # POST handling
    if request.method == 'POST':
        form_type = request.form.get('form_type')

        # 1️⃣ Upload exam file
        if form_type == 'upload':
            time_slot_share = request.form.get("time_slot_share")
            time_slot_open = request.form.get("time_slot_open")
            slot_share_dt = datetime.strptime(time_slot_share, "%Y-%m-%dT%H:%M")
            slot_open_dt = datetime.strptime(time_slot_open, "%Y-%m-%dT%H:%M")

            result = handle_file_upload(
                file_key='exam_file',
                expected_cols=['exam date', 'day', 'start', 'end', 'program', 'course code/section', 'course name', 'lecturer', 'total student by venue', 'venue'],
                process_row_fn=lambda row: process_exam_row(row, slot_share_dt, slot_open_dt),
                redirect_endpoint='admin_manageExam',
                usecols="A:J",
                skiprows=1
            )
            recalc_invigilators_for_new_exams()
            flash("✅ Invigilators recalculated and balanced  for overlapping venue/time slots.", "success")
            return result

        # 2️⃣ Edit / delete exam
        if form_type == "edit" and exam_select:
            action = request.form.get("action")

            if action == "update":
                try:
                    adjust_exam(
                        exam=exam_select,
                        new_start=datetime.fromisoformat(request.form["startDateTime"]),
                        new_end=datetime.fromisoformat(request.form["endDateTime"]),
                        new_venues=request.form.getlist("venue[]"),
                        new_students = [
                            int(s) for s in request.form.getlist("venueStudents[]")
                            if str(s).strip().isdigit()
                        ],
                        time_open=safe_iso(request.form.get("examTimeCreate"), "Open Time"),
                        time_expire=safe_iso(request.form.get("examTimeExpire"), "Expire Time"),
                    )
                    flash(f"💾 Exam [{exam_select.examId} - {exam_select.course.courseName}] updated successfully", "success")
                    record_action("EDIT EXAM", "EXAM", exam_select.examId, user_id)

                except Exception as e:
                    db.session.rollback()
                    flash(str(e), "error")

            elif action == 'delete':
                reset_exam_relations(exam_select)
                flash(f"🗑️ Exam [{exam_select.examId} - {exam_select.course.courseName}] schedule removed", "success")
                record_action("DELETE EXAM", "EXAM", exam_select.examId, user_id)

            return redirect(url_for('admin_manageExam'))

    # Render page
    return render_template(
        'admin/adminManageExam.html',
        active_tab='admin_manageExamtab',
        exam_data=exam_data,
        display_exam_data=display_exam_data,
        unassigned_exam=unassigned_exam,
        total_exam_activated=total_exam_activated,
        edit_exam_data=edit_exam_data,
        exam_select=exam_select,
        venue_data=venue_data,
        department_data=department_data,
    )




# -------------------------------
# Function for Admin User Download Excel Template
# -------------------------------
def generate_user_template():
    warnings.simplefilter("ignore", UserWarning)
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None, "Workbook has no active worksheet"
    ws.title = "Users"

    # First row empty
    ws.append([])
    # Second row = headers
    headers = ['Id', 'CardId', 'Name', 'Department', 'Role', 'Email', 'Contact', 'Gender']
    ws.append(headers)

    # === Hidden sheet for lookup lists ===
    ws_lists = wb.create_sheet(title="Lists")

    # --- Departments ---
    departments = Department.query.all()
    for i, d in enumerate(departments, start=1):
        ws_lists[f"A{i}"] = d.departmentCode

    # --- Roles ---
    roles = ["Lecturer", "Program Officer", "Dean", "HOS", "HOP", "Admin"]
    for i, r in enumerate(roles, start=1):
        ws_lists[f"B{i}"] = r

    # --- Gender ---
    genders = ["Male", "Female"]
    for i, g in enumerate(genders, start=1):
        ws_lists[f"C{i}"] = g

    # === Data Validations ===
    if departments:
        dv_department = DataValidation(
            type="list",
            formula1=f"=Lists!$A$1:$A${len(departments)}",
            allow_blank=False
        )
        ws.add_data_validation(dv_department)
        dv_department.add("D3:D1002")  # Department dropdown

    dv_role = DataValidation(
        type="list",
        formula1=f"=Lists!$B$1:$B${len(roles)}",
        allow_blank=False
    )
    ws.add_data_validation(dv_role)
    dv_role.add("E3:E1002")  # Role dropdown

    dv_gender = DataValidation(
        type="list",
        formula1=f"=Lists!$C$1:$C${len(genders)}",
        allow_blank=False
    )
    ws.add_data_validation(dv_gender)
    dv_gender.add("H3:H1002")  # Gender dropdown

    ws_lists.sheet_state = 'hidden'
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# -------------------------------
# Route to download User Excel template
# -------------------------------
@app.route('/download_user_template')
@login_required
def download_user_template():
    output = generate_user_template()
    return send_file(
        output,
        as_attachment=True,
        download_name="UserTemplate.xlsx",  # type: ignore[arg-type]
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -------------------------------
# Function for Admin ManageStaff Route Upload File, Validate Contact Number
# -------------------------------
def clean_contact(contact):
    if not contact:
        return ""
    contact = str(contact).strip().replace(".0", "")
    contact = "".join(filter(str.isdigit, contact))
    if contact and not contact.startswith("0"):
        contact = "0" + contact
    return contact if 10 <= len(contact) <= 11 else ""

# -------------------------------
# Function for Admin ManageStaff Route Upload File
# -------------------------------
def process_staff_row(row):
    user_id = session.get('user_id')
    hashed_pw = bcrypt.generate_password_hash('Abc12345!').decode('utf-8')

    # Handle empty cardid properly
    cardid = row['cardid']
    if pd.isna(cardid) or str(cardid).strip().lower() in ["", "nan", "none"]:
        cardid = None
    else:
        cardid = str(cardid).upper()

    # Convert gender string to boolean: Male=True, Female=False
    gender_str = str(row['gender']).strip().lower()
    if gender_str == "male":
        gender_bool = True
    elif gender_str == "female":
        gender_bool = False
    else:
        gender_bool = None  # or raise an error if invalid value

    # Clean contact
    contact_cleaned = clean_contact(row['contact'])

    return create_staff(
        userId=user_id,
        id=row['id'],
        department=str(row['department']).upper(),
        name=str(row['name']).upper(),
        role=str(row['role']).upper(),
        email=str(row['email']),
        contact=contact_cleaned,
        gender=gender_bool, 
        hashed_pw=hashed_pw,
        cardId=cardid, 
    )


# -------------------------------
# Read All StaffDetails Under The ManageLecturerEditPage
# -------------------------------
@app.route('/get_staff/<id>')
@login_required
def get_staff(id):
    user = User.query.filter_by(userId=id).first()
    if not user:
        return jsonify({"error": "User not found"}), 404

    return jsonify({
        "userId": user.userId,
        "userName": user.userName,
        "userEmail": user.userEmail,
        "userContact": user.userContact or "",
        "userStatus": str(user.userStatus),
        "userLevel": user.userLevel,
        "userDepartment": user.userDepartment or "",
        "userCardId": user.userCardId or "",
        "userGender": 1 if user.userGender else 0
    })

# -------------------------------
# Function for Admin ManageStaff Route
# -------------------------------
@app.route('/admin/manageStaff', methods=['GET', 'POST'])
@login_required
def admin_manageStaff():
    user_id = session.get('user_id')
    user_data = User.query.order_by(func.field(User.userStatus, 1, 0, 2), User.userLevel.asc(), User.userName.asc()).all()
    department_data = Department.query.all()

    # === Dashboard Counts ===
    total_staff = User.query.count()
    total_admin = User.query.filter_by(userLevel="ADMIN").count()
    total_hop = User.query.filter_by(userLevel="HOP").count()
    total_hos = User.query.filter_by(userLevel="HOS").count()
    total_dean = User.query.filter_by(userLevel="DEAN").count()
    total_lecturer = User.query.filter_by(userLevel="LECTURER").count()
    total_male_staff = User.query.filter_by(userGender=True).count()
    total_female_staff = User.query.filter_by(userGender=False).count()
    total_activated = User.query.filter_by(userStatus=1).count()
    total_deactivate = User.query.filter_by(userStatus=0).count()
    total_deleted = User.query.filter_by(userStatus=2).count()
    # === Staff Department Distribution ===
    staff_dept_query = (
        db.session.query(Department.departmentCode, func.count(User.userId))
        .join(User, User.userDepartment == Department.departmentCode)
        .group_by(Department.departmentCode)
        .all()
    )

    staffDepartmentLabels = [row[0] for row in staff_dept_query]
    staffDepartmentCounts = [row[1] for row in staff_dept_query]

    staff_id = request.form.get('editStaffId')
    user_select = User.query.filter_by(userId=staff_id).first()

    if request.method == 'POST':
        form_type = request.form.get('form_type')

        if form_type == 'upload':
            return handle_file_upload(
                file_key='staff_file',
                expected_cols=['id', 'cardid', 'name', 'department', 'role', 'email', 'contact', 'gender'],
                process_row_fn=process_staff_row,
                redirect_endpoint='admin_manageStaff',
                usecols="A:H",
                skiprows=1 
            )

        elif form_type == 'edit':
            action = request.form.get('action')
            if action == 'update' and user_select:
                user_select.userName = request.form['editUsername']
                user_select.userEmail = request.form['editEmail']
                user_select.userContact = request.form['editContact']
                user_select.userGender = int(request.form['editGender'])
                user_select.userLevel = request.form['editRole']
                user_select.userStatus = int(request.form['editStatus'])
                user_select.userCardId = request.form['editCardId']
                new_department_code = request.form['editDepartment']

                # Always update department and related dean/hop/hos fields
                old_department = Department.query.filter_by(departmentCode=user_select.userDepartment).first()
                new_department = Department.query.filter_by(departmentCode=new_department_code).first()

                # Remove from old department (if user moves OR role changes)
                if old_department:
                    if old_department.hopId == user_select.userId:
                        old_department.hopId = None
                    if old_department.deanId == user_select.userId:
                        old_department.deanId = None
                    if old_department.hosId == user_select.userId:
                        old_department.hosId = None

                # Assign to new department
                if new_department:
                    if user_select.userLevel == "HOS":  # HOS
                        new_department.hosId = user_select.userId
                    elif user_select.userLevel == "HOP":  # HOP
                        new_department.hopId = user_select.userId
                    elif user_select.userLevel == "DEAN":  # Dean
                        new_department.deanId = user_select.userId

                # Update user department & timestamp
                user_select.userDepartment = new_department_code
                
                if user_select.userContact == '':
                    user_select.userContact = None

                db.session.commit()
                flash(f"Staff [{user_select.userId} - {user_select.userName}] updated successfully", "success")
                record_action("EDIT STAFF", "STAFF", user_select.userId, user_id)

            elif action == 'delete' and user_select:
                user_select.userStatus = 2
                db.session.commit() 
                flash("Staff deleted successfully", "success")
                record_action("DELETE STAFF", "STAFF", user_select.userId, user_id)

            return redirect(url_for('admin_manageStaff'))

        elif form_type == 'manual':
            form_data = {
                "userId": user_id,
                "id": request.form.get('userid', '').strip(),
                "department": request.form.get('department', '').strip(),
                "name": request.form.get('username', '').strip(),
                "role": request.form.get('role', '').strip(),
                "email": request.form.get('email', '').strip(),
                "contact": request.form.get('contact', '').strip(),
                "gender": bool(int(request.form.get('gender', 0))), 
                "hashed_pw": bcrypt.generate_password_hash('Abc12345!').decode('utf-8'),
                "cardId": request.form.get('cardId', '').strip(),
            }
            success, message = create_staff(**form_data)
            if success:
                # Send verification email
                email_success, email_message = send_verifyActivateLink(form_data["email"])
                if email_success:
                    flash("Staff account created and verification link sent!", "success")
                else:
                    flash(f"Staff account created but failed to send verification email: {email_message}", "error")
            else:
                flash(message, "error")
            return redirect(url_for('admin_manageStaff'))

    return render_template(
        'admin/adminManageStaff.html',
        active_tab='admin_manageStafftab',
        user_data=user_data,
        department_data=department_data,
        total_staff=total_staff,
        total_admin=total_admin,
        total_hop=total_hop,
        total_hos=total_hos,
        total_dean=total_dean,
        total_lecturer=total_lecturer,
        total_male_staff=total_male_staff,
        total_female_staff=total_female_staff,
        total_activated=total_activated,
        total_deactivate=total_deactivate,
        total_deleted=total_deleted,
        user_select=user_select,
        staffDepartmentLabels=staffDepartmentLabels,
        staffDepartmentCounts=staffDepartmentCounts
    )




# -------------------------------
# Extract Base Name + Timestamp
# -------------------------------
def extract_base_name_and_timestamp(file_name):
    """Extract base name + optional _DDMMYY"""
    pattern = r"^(.*?)(?:_([0-9]{6}))(?:\s.*)?\.pdf$"
    match = re.match(pattern, file_name, re.IGNORECASE)

    if match:
        base_name = match.group(1).strip()
        timestamp_str = match.group(2)
        try:
            timestamp = datetime.strptime(timestamp_str, "%d%m%y")
        except ValueError:
            timestamp = None
    else:
        base_name = re.sub(r"(_\d{6}.*)?\.pdf$", "", file_name, flags=re.IGNORECASE).strip()
        timestamp = None

    return base_name, timestamp

# -------------------------------
# Parse Activity Line
# -------------------------------
def parse_activity(line):
    activity = {}

    m_type = re.match(r"(LECTURE|TUTORIAL|PRACTICAL)", line)
    if m_type:
        activity["class_type"] = m_type.group(1)

    m_time = re.search(r",(\d{2}:\d{2}-\d{2}:\d{2})", line)
    if m_time:
        activity["time"] = m_time.group(1)

    m_weeks = re.search(r"WEEKS:([^C]+)", line)
    if m_weeks:
        weeks_data = m_weeks.group(1).split(",")
        if len(weeks_data) > 1:
            activity["weeks_range"] = weeks_data[:-1]
            activity["weeks_date"]  = weeks_data[-1]
        else:
            activity["weeks_range"] = weeks_data

    m_course = re.search(r"COURSES:([^;]+);", line)
    if m_course:
        activity["course"] = m_course.group(1)

    m_sections = re.search(r"SECTIONS:(.+?)ROOMS", line)
    if m_sections:
        sections = m_sections.group(1).strip(";").split(";")
        activity["sections"] = []
        for sec in sections:
            if "|" in sec:
                intake, code, sec_name = sec.split("|")
                activity["sections"].append({
                    "intake": intake,
                    "course_code": code,
                    "section": sec_name
                })

    m_room = re.search(r"ROOMS:([^;]+);", line)
    if m_room:
        activity["room"] = m_room.group(1)

    return activity

# -------------------------------
# Parse Timetable Text
# -------------------------------
def parse_timetable(raw_text):
    text_no_whitespace = re.sub(r"\s+", "", raw_text)

    lecturer_name = "UNKNOWN"
    timerow = ""

    # Try extract title/timerow
    title_match = re.match(r"^(.*?)(07:00.*?23:00)", text_no_whitespace)
    if title_match:
        title_raw   = title_match.group(1)
        timerow     = title_match.group(2)
    else:
        title_raw = ""

    # Extract lecturer name
    try:
        name_match = re.search(r"-([^-()]+)\(", title_raw)
        if name_match:
            raw_name            = name_match.group(1)
            formatted_name      = re.sub(r'(?<!^)([A-Z])', r' \1', raw_name).strip()
            formatted_name      = formatted_name.replace("A/ P", "A/P").replace("A/ L", "A/L")
            lecturer_name       = formatted_name
            text_no_whitespace  = text_no_whitespace.replace(raw_name, lecturer_name.replace(" ", ""))
    except Exception:
        pass

    text = text_no_whitespace.upper()
    match_title = re.match(r"^(.*?)(07:00.*?23:00)", text)
    if match_title:
        title = match_title.group(1).strip()
        timerow = match_title.group(2).strip()
        text = text.replace(title, "").replace(timerow, "")
    else:
        title = "TIMETABLE"

    days = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    for day in days:
        text = re.sub(day, f"\n\n{day}", text, flags=re.IGNORECASE)

    for kw in ["LECTURE", "TUTORIAL", "PRACTICAL", "PUBLISHED"]:
        sep = "\n\n" if kw == "PUBLISHED" else "\n"
        text = re.sub(kw, f"{sep}{kw}", text, flags=re.IGNORECASE)

    text = re.sub(r"\n{3,}", "\n\n", text)

    structured = {
        "title"     : title,
        "lecturer"  : lecturer_name,
        "timerow"   : timerow,
        "days"      : {}
    }

    current_day = None
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if line in days:
            current_day = line
            structured["days"][current_day] = []
        elif current_day and any(kw in line for kw in ["LECTURE", "TUTORIAL", "PRACTICAL"]):
            structured["days"][current_day].append(parse_activity(line))

    return structured


def parse_date_range(date_range):
    """Parse classWeekDate 'MM/DD/YYYY-MM/DD/YYYY' and return (start, end) datetime safely."""
    if not date_range:
        return None, None
    try:
        # Split only on the last hyphen in case there are others in the string
        start_str, end_str = date_range.rsplit("-", 1)
        start = datetime.strptime(start_str.strip(), "%m/%d/%Y")
        end = datetime.strptime(end_str.strip(), "%m/%d/%Y")
        return start, end
    except Exception as e:
        print(f"Date parse error for '{date_range}': {e}")
        return None, None


# -------------------------------
# Get TimetableLink Details for ManageTimetableEditPage
# -------------------------------
@app.route('/get_linkTimetable/<path:timetableID>')
@login_required
def get_linkTimetable(timetableID):
    timetable = Timetable.query.filter_by(timetableId=timetableID).first()
    if not timetable:
        return jsonify({"error": "Timetable not found"}), 404
    return jsonify({
        "timetableId": timetable.timetableId,
        "user_id": timetable.user_id
    })

# -------------------------------
# Save Parsed Timetable to DB
# -------------------------------
def save_timetable_to_db(structured):
    lecturer = structured.get("lecturer")
    filename = structured.get("filename")

    if not lecturer:
        return

    # Normalize lecturer name (remove all spaces)
    normalized_lecturer = ''.join(lecturer.split())

    # Find user where username (spaces removed) matches normalized lecturer
    user = User.query.filter(func.replace(User.userName, " ", "") == normalized_lecturer).first()

    # Ensure timetable exists for user
    timetable = None
    if user:
        timetable = Timetable.query.filter_by(user_id=user.userId).first()
        if not timetable:
            timetable = Timetable(user_id=user.userId)
            db.session.add(timetable)
            db.session.commit()

    # ---- Delete old rows if lecturer already exists in DB ----
    existing_rows = TimetableRow.query.filter_by(lecturerName=lecturer).count()
    if existing_rows > 0:
        TimetableRow.query.filter_by(lecturerName=lecturer).delete()
        db.session.commit()

    rows_inserted = 0

    # ---- Insert new timetable rows ----
    for day, activities in structured.get("days", {}).items():
        for act in activities:
            class_type = act.get("class_type", "").strip()
            course_name_uploaded = act.get("course", "").strip()
            if not (class_type and act.get("time") and act.get("room") and course_name_uploaded):
                continue

            # Normalize course name for comparison
            normalized_course_uploaded = ''.join(course_name_uploaded.split()).lower()

            # Try to find matching course in database (ignoring spaces)
            matching_course = None
            for course in Course.query.all():
                normalized_course_db = ''.join(course.courseName.split()).lower()
                if normalized_course_uploaded == normalized_course_db:
                    matching_course = course
                    break

            # If found and user exists, assign lecturer to tutorial/practical
            if matching_course and user:
                if class_type.lower() == "tutorial":
                    matching_course.courseTutorial = user.userId
                elif class_type.lower() == "practical":
                    matching_course.coursePractical = user.userId
                db.session.commit()

            # Insert timetable rows (if have sections)
            if act.get("sections"):
                for sec in act["sections"]:
                    if not (sec.get("intake") and sec.get("course_code") and sec.get("section")):
                        continue

                    new_row = TimetableRow(
                        timetable_id=timetable.timetableId if timetable else None,
                        filename=filename,
                        lecturerName=lecturer,
                        classType=class_type,
                        classDay=day,
                        classTime=act.get("time"),
                        classRoom=act.get("room"),
                        courseName=course_name_uploaded,
                        courseIntake=sec.get("intake"),
                        courseCode=sec.get("course_code"),
                        courseSection=sec.get("section"),
                        classWeekRange=",".join(act.get("weeks_range", [])) if act.get("weeks_range") else None,
                        classWeekDate=act.get("weeks_date"),
                    )
                    db.session.add(new_row)
                    rows_inserted += 1

    db.session.commit()
    return rows_inserted



# -------------------------------
# Function for Admin ManageTimetable Route
# -------------------------------
@app.route('/admin/manageTimetable', methods=['GET', 'POST'])
@login_required
def admin_manageTimetable():
    department_data = Department.query.all()
    selected_department = request.args.get("department")
    selected_lecturer = request.args.get("lecturer")

    # Base query
    timetable_data_query = (
        TimetableRow.query
        .join(Timetable, TimetableRow.timetable_id == Timetable.timetableId)
        .join(User, Timetable.user_id == User.userId)
        .filter(User.userStatus == 1)  # active staff only
    )

    # Department filter
    if selected_department:
        timetable_data_query = timetable_data_query.filter(User.userDepartment == selected_department)
    # Lecturer filter (by userId)
    if selected_lecturer:
        timetable_data_query = timetable_data_query.filter(TimetableRow.lecturerName == selected_lecturer)


    timetable_data = timetable_data_query.order_by(TimetableRow.rowId.asc()).all()
    lecturers = sorted({row.lecturerName for row in timetable_data})
    total_timetable = db.session.query(func.count(func.distinct(TimetableRow.lecturerName))).scalar()
    timetable_list = Timetable.query.filter(Timetable.timetableId != None).all()
    
    timetable_select = None
    timetable_selected = request.form.get('editTimetableList')
    if timetable_selected:
        timetable_select = Timetable.query.filter_by(timetableId=timetable_selected).first()
    timetable_map = {t.user_id: t.timetableId for t in timetable_list}

    # Staff list (exclude certain levels/status)
    staff_all  = User.query.filter(
        User.userLevel != 5,
        User.userStatus != 2
    ).all()
    unassigned_staff_list = [staff for staff in staff_all if staff.userId not in timetable_map]
    staff_list = staff_all

    # Count timetable per day
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    day_counts = {
        f"{day.lower()}_timetable": db.session.query(TimetableRow.courseCode)
            .filter(TimetableRow.classDay == day).distinct().count()
        for day in days
    }

    # Group all lecturers with unassigned rows
    grouped_unassigned = defaultdict(int)
    for row in TimetableRow.query.filter_by(timetable_id=None).all():
        grouped_unassigned[row.lecturerName] += 1

    unassigned_summary = [{"lecturer": name, "count": count} for name, count in grouped_unassigned.items()]

    # ---- POST Handling ----
    if request.method == "POST":
        form_type = request.form.get('form_type')

        # --- Upload timetable PDF files ---
        if form_type == 'upload':
            files = request.files.getlist("timetable_file[]")
            latest_files = {}
            skipped_files = []

            # Keep only the latest file per lecturer
            for file in files:
                base_name, timestamp = extract_base_name_and_timestamp(file.filename)
                if not base_name:
                    continue
                if base_name not in latest_files:
                    latest_files[base_name] = (timestamp, file)
                else:
                    existing_timestamp, existing_file = latest_files[base_name]
                    if timestamp and (existing_timestamp is None or timestamp > existing_timestamp):
                        skipped_files.append(existing_file.filename)
                        latest_files[base_name] = (timestamp, file)
                    else:
                        skipped_files.append(file.filename)

            total_rows_inserted = 0
            total_files_processed = 0

            # Process each latest file
            for base_name, (timestamp, file) in latest_files.items():
                reader = PyPDF2.PdfReader(file.stream)
                raw_text = "".join(page.extract_text() + " " for page in reader.pages if page.extract_text())
                structured = parse_timetable(raw_text)
                structured['filename'] = file.filename
                rows_inserted = save_timetable_to_db(structured)

                if (rows_inserted or 0) > 0:
                    total_files_processed += 1
                    total_rows_inserted += rows_inserted or 0
            flash(f"Files read: {len(files)}, Processed: {total_files_processed}, Rows inserted: {total_rows_inserted}, Files skipped: {len(skipped_files)}", "success")
            return redirect(url_for('admin_manageTimetable'))

        # --- Manual link timetable to staff ---
        elif form_type == 'manual':
            user_id = request.form.get("staffList")
            lecturer = request.form.get("lecturerName")

            if user_id and lecturer:
                timetable = Timetable.query.filter_by(user_id=user_id).first()
                if not timetable:
                    timetable = Timetable(user_id=user_id)
                    db.session.add(timetable)
                    db.session.commit()

                rows = TimetableRow.query.filter_by(lecturerName=lecturer, timetable_id=None).all()
                for row in rows:
                    row.timetable_id = timetable.timetableId

                db.session.commit()
                flash(f"Timetable for {lecturer} has been successfully linked to Staff ID {user_id}.", "success")
            else:
                flash("Missing lecturer or staff", "error")
            return redirect(url_for('admin_manageTimetable'))

        # --- Edit / Delete timetable link ---
        elif form_type == 'edit':
            action = request.form.get('action')
            if action == 'update' and timetable_select:
                new_user_id = request.form['editStaffList']

                if str(timetable_select.user_id) == new_user_id:
                    flash("No changes made. Timetable already linked to this staff.", "success")
                else:
                    existing = Timetable.query.filter(
                        Timetable.user_id == new_user_id,
                        Timetable.timetableId != timetable_select.timetableId
                    ).first()

                    if existing:
                        flash(f"Staff ID:{new_user_id} is already linked to another timetable(ID:{existing.timetableId}).", "error")
                    else:
                        timetable_select.user_id = new_user_id
                        db.session.commit()
                        flash("Timetable updated successfully.", "success")

            elif action == 'delete' and timetable_select:
                db.session.delete(timetable_select)
                db.session.commit()
                flash("Timetable deleted successfully.", "success")
                return redirect(url_for('admin_manageTimetable'))

    return render_template('admin/adminManageTimetable.html',active_tab='admin_manageTimetabletab',timetable_data=timetable_data,lecturers=lecturers,selected_lecturer=selected_lecturer,
        department_data=department_data,selected_department=selected_department,unassigned_staff_list=unassigned_staff_list,total_timetable=total_timetable,unassigned_summary=unassigned_summary,staff_list=staff_list,
        **day_counts,timetable_list=timetable_list,timetable_map=timetable_map,timetable_select=timetable_select)





@app.route("/admin/updateAttendanceTime", methods=["POST"])
@login_required
def update_attendance_time():
    data = request.get_json()
    attendance_id = data.get("attendance_id")
    check_in_str = data.get("check_in")
    check_out_str = data.get("check_out")
    invigilation_status = data.get("invigilation_status")  # Get the new status

    # Handle empty values
    check_in_str = None if not check_in_str or check_in_str == "None" else check_in_str
    check_out_str = None if not check_out_str or check_out_str == "None" else check_out_str

    att = InvigilatorAttendance.query.get(attendance_id)
    if not att:
        return jsonify({"success": False, "message": "Attendance not found."}), 404

    # Parse datetime
    def parse_datetime(dt_str):
        if not dt_str: return None
        try:
            return datetime.strptime(dt_str, "%Y-%m-%dT%H:%M:%S")
        except ValueError:
            return datetime.strptime(dt_str, "%Y-%m-%dT%H:%M")

    try:
        check_in = parse_datetime(check_in_str)
        check_out = parse_datetime(check_out_str)
    except Exception as e:
        return jsonify({"success": False, "message": f"Invalid datetime format: {str(e)}"}), 400

    exam = att.report.exam
    exam_start = exam.examStartTime
    exam_end = exam.examEndTime
    allowed_start = exam_start - timedelta(hours=1)
    allowed_end = exam_end + timedelta(hours=1)

    # Validate times
    if check_in and check_out:
        if not (allowed_start <= check_in <= allowed_end and allowed_start <= check_out <= allowed_end):
            return jsonify({"success": False, "message": "Time must be within 1 hour before/after exam period."}), 400
        if check_in >= check_out:
            return jsonify({"success": False, "message": "Check-in must be before check-out."}), 400

    def calculate_hours(start, end):
        """Calculate overlapping hours between exam and attendance."""
        if not start or not end:
            return 0.0
        adj_start = max(start, exam_start)
        adj_end = min(end, exam_end)
        if adj_start >= adj_end:
            return 0.0
        return round((adj_end - adj_start).total_seconds() / 3600.0, 2)

    # --- Hours before & after update ---
    old_hours = calculate_hours(att.checkIn, att.checkOut)
    new_hours = calculate_hours(check_in, check_out)
    exam_hours = round((exam_end - exam_start).total_seconds() / 3600.0, 2)

    # Update invigilation status if provided
    if invigilation_status is not None:
        att.invigilationStatus = invigilation_status
        att.timeAction = datetime.now() + timedelta(hours=8)    

    invigilator = att.invigilator

    # --- Safely update cumulative hours and pending hours ---
    if att.invigilationStatus:
        # Remove previous contribution
        if att.checkIn and att.checkOut:
            invigilator.userCumulativeHours -= old_hours
            invigilator.userPendingCumulativeHours += exam_hours  # restore previous pending

        # Add new contribution
        if check_in and check_out:
            invigilator.userCumulativeHours += new_hours
            invigilator.userPendingCumulativeHours -= exam_hours  # deduct for this completed exam

    # --- Determine remark ---
    remark = "PENDING"
    if check_in:
        remark = "CHECK IN LATE" if check_in > exam_start else "CHECK IN"
        if check_out:
            if check_out < exam_end:
                remark = "CHECK IN LATE" if "LATE" in remark else "CHECK OUT EARLY"
            elif check_out <= allowed_end:
                remark = "COMPLETED" if "LATE" not in remark else remark
            else:
                remark = "EXPIRED"

    # --- Apply updates ---
    att.checkIn = check_in
    att.checkOut = check_out
    att.remark = remark

    db.session.commit()

    return jsonify({
        "success": True,
        "message": "Attendance updated successfully.",
        "remark": remark,
        "new_hours": round(new_hours, 2),
        "invigilation_status": att.invigilationStatus,
        "check_in": check_in.strftime("%d/%b/%Y %H:%M:%S") if check_in else "None",
        "check_out": check_out.strftime("%d/%b/%Y %H:%M:%S") if check_out else "None"
    })





# -------------------------------
# Function for Admin ManageInviglationTimetable Route (Simple Calendar View + Overnight Handling)
# -------------------------------
def get_calendar_data():
    venue_exams = (VenueExam.query.join(Exam).all())
    calendar_data = defaultdict(lambda: defaultdict(list))

    for ve in venue_exams:
        exam = ve.exam
        start_dt = ve.startDateTime
        end_dt = ve.endDateTime
        date_key = start_dt.date()

        calendar_data[date_key][ve.venueNumber].append({
            "exam_id": exam.examId,
            "course_code": exam.course.courseCodeSectionIntake,
            "course_name": exam.course.courseName,
            "start_time": start_dt,
            "end_time": end_dt,
            "capacity": ve.capacity,
            "is_overnight": start_dt.date() != end_dt.date(),
        })

    return dict(sorted(calendar_data.items()))


# -------------------------------
# Function for Admin ManageInviglationTimetable Route
# -------------------------------
@app.route('/admin/manageInvigilationTimetable', methods=['GET'])
@login_required
def admin_manageInvigilationTimetable():
    calendar_data = get_calendar_data()
    return render_template('admin/adminManageInvigilationTimetable.html', active_tab='admin_manageInvigilationTimetabletab', calendar_data=calendar_data)



# -------------------------------
# Calculate All InvigilatorAttendance and InvigilationReport Data From Database
# -------------------------------
def calculate_invigilation_stats():
    base_query = (
        db.session.query(InvigilationReport)
        .join(Exam, InvigilationReport.examId == Exam.examId)
        .join(InvigilatorAttendance, InvigilatorAttendance.reportId == InvigilationReport.invigilationReportId)
        .distinct()
    )

    total_report = base_query.count()
    total_active_report = base_query.filter(Exam.examStatus == True).count()

    # Pull all attendance records for detailed time analysis
    query = (
        db.session.query(
            InvigilatorAttendance.attendanceId,
            InvigilatorAttendance.invigilatorId,
            InvigilatorAttendance.checkIn,
            InvigilatorAttendance.checkOut,
            Exam.examStartTime,
            Exam.examEndTime,
            InvigilatorAttendance.reportId
        )
        .join(InvigilationReport, InvigilatorAttendance.reportId == InvigilationReport.invigilationReportId)
        .join(Exam, InvigilationReport.examId == Exam.examId)
        .filter(InvigilatorAttendance.invigilationStatus == True)
        .all()
    )

    stats = {
        "total_report": total_report,
        "total_activeReport": total_active_report,
        "total_checkInLate": 0,
        "total_checkOutEarly": 0,
    }

    for row in query:
        if row.checkIn and row.checkIn > row.examStartTime:
            stats["total_checkInLate"] += 1
        if row.checkOut and row.checkOut < row.examEndTime:
            stats["total_checkOutEarly"] += 1
    return stats


# -------------------------------
# Read All InvigilatorAttendance Data From Database
# -------------------------------
def get_all_attendances():
    return (
        InvigilatorAttendance.query
        .join(InvigilationReport, InvigilatorAttendance.reportId == InvigilationReport.invigilationReportId)
        .join(Exam, InvigilationReport.examId == Exam.examId)
        .order_by(
            Exam.examStatus.desc(), 
            InvigilatorAttendance.invigilationStatus.desc(),
            InvigilatorAttendance.rejectReason.is_(None).desc())
        .all()
    )


# -------------------------------
# Helper: Parse Date + Time from Excel
# -------------------------------
def parse_attendance_datetime(date_val, time_val):
    """Combine date + time from Excel into datetime object."""
    try:
        # Parse date
        if isinstance(date_val, datetime):
            date_part = date_val.date()
        else:
            date_str = str(date_val).strip()
            try:
                # Try DD/MM/YYYY first
                date_part = datetime.strptime(date_str, "%d/%m/%Y").date()
            except ValueError:
                # Try ISO format YYYY-MM-DD
                date_part = datetime.strptime(date_str.split()[0], "%Y-%m-%d").date()
        
        # Parse time
        if isinstance(time_val, datetime):
            time_part = time_val.time()
        else:
            time_str = str(time_val).strip()
            if ':' in time_str:
                parts = time_str.split(":")
                if len(parts) == 2:
                    hours = int(parts[0])
                    minutes = int(parts[1])
                    time_part = datetime.strptime(f"{hours:02d}:{minutes:02d}:00", "%H:%M:%S").time()
                elif len(parts) == 3:
                    time_part = datetime.strptime(time_str, "%H:%M:%S").time()
                else:
                    raise ValueError(f"Invalid time format: {time_str}")
            else:
                raise ValueError(f"Invalid time format: {time_str}")
        
        return datetime.combine(date_part, time_part)
    
    except Exception as e:
        flash(f"Error parsing datetime: {e}", "error")
        return None


# -------------------------------
# Process Single Attendance Row
# -------------------------------
def process_attendance_row(row):
    try:
        # 1. Extract UID
        raw_uid = str(row['card iud']).upper().replace('UID:', '').strip().replace(' ', '')

        # 2. Find matching user
        user = User.query.filter_by(userCardId=raw_uid).first()
        if not user:
            return False, f"No matching user for UID {raw_uid}"

        # 3. Parse datetime
        dt_obj = parse_attendance_datetime(row['date'], row['time'])
        if not dt_obj:
            return False, f"Invalid date/time format: {row['date']} {row['time']}"

        # 4. Determine in/out value
        inout_val = str(row['in/out']).lower().strip()
        if inout_val not in ['in', 'out']:
            return False, f"Invalid in/out value: {row['in/out']}"

        # 5. Skip duplicate
        if inout_val == 'in':
            existing = InvigilatorAttendance.query.filter_by(
                invigilatorId=user.userId,
                checkIn=dt_obj
            ).first()
        else:
            existing = InvigilatorAttendance.query.filter_by(
                invigilatorId=user.userId,
                checkOut=dt_obj
            ).first()

        if existing:
            return False, f"Duplicate entry skipped for {user.userName} at {dt_obj} ({inout_val.upper()})"

        # 6. Proceed with updating attendance
        attendances = InvigilatorAttendance.query \
            .join(InvigilationReport, InvigilatorAttendance.reportId == InvigilationReport.invigilationReportId) \
            .join(Exam, InvigilationReport.examId == Exam.examId) \
            .filter(InvigilatorAttendance.invigilatorId == user.userId) \
            .all()
        
        if not attendances:
            return False, f"No invigilation sessions found for user {user.userName} near {dt_obj}"

        updated_count = 0
        for attendance in attendances:
            exam = Exam.query.get(attendance.report.examId)
            if not exam:
                continue

            exam_start = exam.examStartTime
            exam_end = exam.examEndTime

            if not (exam_start - timedelta(hours=1) <= dt_obj <= exam_end + timedelta(hours=1)):
                continue

            if inout_val == "in":
                if attendance.checkIn is None:
                    attendance.checkIn = dt_obj
                    attendance.remark = "CHECK IN" if dt_obj <= exam_start else "CHECK IN LATE"
                    updated_count += 1
            elif inout_val == "out":
                if attendance.checkOut is None:
                    attendance.checkOut = dt_obj
                    if attendance.checkIn:
                        actual_checkin = attendance.checkIn
                        checkin_for_hours = actual_checkin if attendance.remark == "CHECK IN LATE" else exam_start
                        checkout_for_hours = dt_obj if dt_obj < exam_end else exam_end
                        hours_worked = (checkout_for_hours - checkin_for_hours).total_seconds() / 3600
                        user.userCumulativeHours += hours_worked
                        attendance.remark = "COMPLETED" if dt_obj >= exam_end else "CHECK OUT EARLY"
                    else:
                        attendance.remark = "PENDING"
                    updated_count += 1

        if updated_count > 0:
            db.session.commit()
            return True, f"Attendance updated for {user.userName} ({updated_count} record(s))"
        else:
            return False, f"No attendance sessions updated for {user.userName} at {dt_obj}"

    except Exception as e:
        db.session.rollback()
        return False, f"Error processing row: {e}"


@app.route('/get_report/<int:report_id>')
@login_required
def get_report(report_id):
    report = InvigilationReport.query.get(report_id)
    if not report:
        return jsonify({"error": "Invigilation report not found"}), 404
    
    exam = report.exam
    course = exam.course

    attendances = []
    for att in report.attendances:
        attendances.append({
            "attendanceId": att.attendanceId,
            "invigilatorId": att.invigilatorId,
            "invigilatorName": att.invigilator.userName,
            "gender": att.invigilator.userGender,
            "venue": att.venueNumber
        })

    return jsonify({
        "examId": exam.examId,
        "courseCode": course.courseCodeSectionIntake,
        "courseName": course.courseName,
        "examStart": exam.examStartTime.strftime("%Y-%m-%d %H:%M"),
        "examEnd": exam.examEndTime.strftime("%Y-%m-%d %H:%M"),
        "attendances": attendances
    })

@app.route('/get_valid_invigilators')
@login_required
def get_valid_invigilators():
    valid = User.query.filter_by(userLevel=1).all()
    return jsonify([{"userId": u.userId, "userName": u.userName}for u in valid])

# -------------------------------
# Admin Route
# -------------------------------
@app.route('/admin/manageInvigilationReport', methods=['GET', 'POST'])
@login_required
def admin_manageInvigilationReport():
    reports = (
        InvigilationReport.query
        .join(Exam, InvigilationReport.examId == Exam.examId)
        .filter(Exam.examStatus==True)
        .all()
    )
    
    attendances = get_all_attendances()
    stats = calculate_invigilation_stats()

    # Attach composite key for sorting/grouping
    for att in attendances:
        report = att.report
        exam = Exam.query.get(report.examId) if report else None
        att.group_key = (
            not exam.examStatus if exam else True,
            exam.examStartTime if exam else datetime.min,
            exam.examId if exam else 0
        )

    if request.method == 'POST':
        form_type = request.form.get('form_type')

        if form_type == 'upload':
            return handle_file_upload(
                file_key='attendance_file',
                expected_cols=['card iud', 'name', 'date', 'time', 'in/out'],
                process_row_fn=process_attendance_row,
                redirect_endpoint='admin_manageInvigilationReport',
                usecols="A:E",
                skiprows=0
            )

        elif form_type == 'edit':
            report_id = request.form.get('reportId')
            report = InvigilationReport.query.get(report_id)
            if not report:
                flash("Report not found.", "error")
                return redirect(url_for('admin_manageInvigilationReport'))

            exam = report.exam
            exam_duration = (exam.examEndTime - exam.examStartTime).total_seconds() / 3600
            selected_invigilators = []

            # First pass: validation
            for key, value in request.form.items():
                if key.startswith("slot_"):
                    invigilator_id = int(value)

                    # Duplicate check within this report
                    if invigilator_id in selected_invigilators:
                        flash(f"Invigilator assigned to multiple slots in the same report.", "error")
                        return redirect(url_for('admin_manageInvigilationReport'))
                    selected_invigilators.append(invigilator_id)

                    # 30-minute gap check
                    att_id = key.replace("slot_", "")
                    att = InvigilatorAttendance.query.get(att_id)
                    invig_attendances = InvigilatorAttendance.query.join(InvigilationReport).join(Exam)\
                        .filter(InvigilatorAttendance.invigilatorId == invigilator_id).all()

                    for ia in invig_attendances:
                        if ia.report.examId == exam.examId:
                            continue  # skip same exam
                        gap_start = ia.report.exam.examStartTime - timedelta(minutes=30)
                        gap_end = ia.report.exam.examEndTime + timedelta(minutes=30)
                        if exam.examStartTime < gap_end and exam.examEndTime > gap_start:
                            flash(f"Invigilator {ia.invigilator.userName} does not have enough gap between exams.", "error")
                            return redirect(url_for('admin_manageInvigilationReport'))

            # Second pass: update records
            for key, value in request.form.items():
                if key.startswith("slot_"):
                    attendance_id = key.replace("slot_", "")
                    new_invigilator_id = int(value)
                    att = InvigilatorAttendance.query.get(attendance_id)
                    old_invigilator = User.query.get(att.invigilatorId)
                    new_invigilator = User.query.get(new_invigilator_id)

                    if old_invigilator.userId != new_invigilator.userId:
                        old_invigilator.userPendingCumulativeHours = max(0, old_invigilator.userPendingCumulativeHours - exam_duration)
                        # Add hours to new invigilator
                        new_invigilator.userPendingCumulativeHours += exam_duration
                        # Update the attendance
                        att.invigilatorId = new_invigilator_id
            db.session.commit()
            flash("Invigilators updated successfully.", "success")
            return redirect(url_for('admin_manageInvigilationReport'))
    return render_template('admin/adminManageInvigilationReport.html', active_tab='admin_manageInvigilationReporttab', attendances=attendances, **stats, reports=reports)

# -------------------------------
# Function for Admin ManageProfile Route
# -------------------------------
@app.route('/admin/profile', methods=['GET', 'POST'])
@login_required
def admin_profile():
    adminId = session.get('user_id')
    admin = User.query.filter_by(userId=adminId).first()

    # Default values for GET requests
    admin_cardUID = admin.userCardId or ''
    admin_contact_text = admin.userContact or ''
    admin_password1_text = ''
    admin_password2_text = ''
    
    # --------------------- MANUAL EDIT PROFILE FORM ---------------------
    if request.method == 'POST':
        admin_cardUID = request.form.get('cardUID', '').strip().replace(' ', '')
        admin_contact_text = request.form.get('contact', '').strip()
        admin_password1_text = request.form.get('password1', '').strip()
        admin_password2_text = request.form.get('password2', '').strip()

        valid, message = check_profile(adminId, admin_cardUID, admin_contact_text, admin_password1_text, admin_password2_text)
        if not valid:
            flash(message, 'error')
            return redirect(url_for('admin_profile'))

        if valid and admin:
            admin.userContact = admin_contact_text or None
            admin.userCardId = admin_cardUID or None
            # Update password only if entered
            if admin_password1_text:
                hashed_pw = bcrypt.generate_password_hash(admin_password1_text).decode('utf-8')
                admin.userPassword = hashed_pw

            db.session.commit()
            flash("Successfully updated", 'success')
            record_action("UPDATE", "PROFILE", adminId, adminId)
            return redirect(url_for('admin_profile'))

    return render_template('admin/adminProfile.html', active_tab='admin_profiletab', admin_data=admin, admin_contact_text=admin_contact_text, 
                           admin_password1_text=admin_password1_text, admin_password2_text=admin_password2_text, admin_cardUID=admin_cardUID)


# -------------------------------
# Function for Admin ViewActivity Route
# -------------------------------
@app.route('/admin/activity', methods=['GET', 'POST'])
@login_required
def admin_activity():
    record = Action.query.order_by(desc(Action.actionTime)).all()
    return render_template('admin/adminActivity.html', active_tab='admin_activitytab', record=record)


